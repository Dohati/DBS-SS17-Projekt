# -*- coding: utf-8 -*-
import psycopg2
from openpyxl import Workbook
from openpyxl import load_workbook

def connect():
	conn=None
	try:
		print('Connecting to the PostgreSQL database...')
		conn = psycopg2.connect(host="localhost", database="Election", user="postgres", password="postgres")

		cur = conn.cursor()

		cur.execute('DROP TABLE IF EXISTS "tweet" CASCADE;')

		cur.execute('DROP TABLE IF EXISTS "hashtag" CASCADE;')

		cur.execute('DROP TABLE IF EXISTS "enthält" CASCADE;')



		cur.execute('CREATE TABLE "tweet" (id integer PRIMARY KEY, handle text NOT NULL, text text NOT NULL, time text NOT NULL, retweet_count integer DEFAULT 0, favorite_count integer DEFAULT 0);')

		cur.execute('CREATE TABLE "hashtag" (id integer PRIMARY KEY, name text NOT NULL UNIQUE);')

		cur.execute('CREATE TABLE "enthält" (tweet_id integer REFERENCES tweet (iD), hashtag_id integer REFERENCES hashtag (iD));')

		conn.commit()

		#öffne bearbeitete datei mit Daten und erstelle Einträge für die BD Election in der Form:
		#"INSERT INTO tweet (id, handle, text, retweet_count, favorite_count, time) VALUES (%s, %s)"

		nDB = load_workbook('neueDatenbank.xlsx') #Neue Datenbank wird geladen

		wsnDB = nDB.active #Aktives Worksheet der gegebenen Datenbank

		row_count = wsnDB.max_row

#Füllung von tweet; Einfach auslesen und schreiben in die Datenbank
		for i in range(2, row_count):
			handle = wsnDB['A'+str(i)].value
			text = wsnDB['B'+str(i)].value 
			time = wsnDB['C'+str(i)].value
			retweet = wsnDB['D'+str(i)].value
			favorite = wsnDB['E'+str(i)].value
			cur.execute("INSERT INTO tweet (id, handle, text, time, retweet_count, favorite_count) VALUES (%s, %s, %s, %s, %s, %s)", (i-1, handle, text, time, int(retweet), int(favorite)))

#Vorbereitung für die Füllung von Hashtag
		liste = [] #leere Liste zum Zwischenspeichern aller Hashtags
		for i in range(2, row_count):
			hashs = wsnDB['F'+str(i)].value #Zelle 'Fi' auslesen
			if (type(hashs) is not unicode): #Überprüfung, ob die Zelle leer ist
				continue #wenn leer, dann weiter mit nächster Zelle
			temp = '#' #temporäre Variable zum Speichern eines Hashtags
			for j in range(1, len(hashs)): #iteration über den unicode string ab Position 1, da # schon gespeichert in temp gespeichert
				if (j == len(hashs)-1): #wenn das letzte Zeichen erreicht
					temp += hashs[j] #Speicher es in temp
					if temp not in liste: #überprüfe ob temp schon in der liste, damit doppelte Einträge vermieden werden
						liste.append(temp) #hänge an liste an
					break
				if (hashs[j]=='#'): #wenn nächstes zeichen ein #
					if temp not in liste: #überprüfe ob temp schon in der liste, damit doppelte Einträge vermieden werden
						liste.append(temp) #hänge an liste an
					temp = '#' #speicher # in temp für nächstes hashtag
					continue #und mach weiter mit nächstem Zeichen
				else: #für alle anderen Fälle
					temp += hashs[j] #hänge an temp den aktuellen Buchstaben

#Füllung von Hashtag
		for i in range(0, len(liste)): #duchrlaufe liste
			cur.execute("INSERT INTO hashtag (id, name) VALUES (%s, %s)", (i+1, liste[i])) #für jeden Eintrag in Liste erstelle einen neuen Eintrag in der sql-Tabelle hashtag

#Füllung von enthält
		for i in range(2, row_count):
			hashs = wsnDB['F'+str(i)].value #Zelle 'Fi' auslesen
			if (type(hashs) is not unicode): #Überprüfung, ob die Zelle leer ist
				continue #wenn leer, dann weiter mit nächster Zelle
			temp = '#' #temporäre Variable zum Speichern eines Hashtags
			for j in range(1, len(hashs)): #iteration über den unicode string ab Position 1, da # schon gespeichert in temp
				if (j == len(hashs)-1): #wenn das letzte Zeichen erreicht
					temp += hashs[j] #dann Speicher es in temp
					cur.execute('SELECT id FROM hashtag WHERE name=%s', (temp,)) #execute eine select Anfrage, um die id des Hashtags zu bekommen
					row = cur.fetchall() #hole das Ergebniss
					hash_id = row[0][0] #speicher das Ergebnis in Variable
					cur.execute("INSERT INTO enthält (tweet_id, hashtag_id) VALUES (%s, %s)", (i-1, hash_id)) #erstelle Datenbankeintrag für Tabelle enthält mit tweet_id und hash-id
					break #beende for-Schleife
				if (hashs[j]=='#'): #wenn nächstes zeichen ein #
					cur.execute('SELECT id FROM hashtag WHERE name=%s', (temp,)) #execute eine select Anfrage, um die id des Hashtags zu bekommen
					row = cur.fetchall() #hole das Ergebniss
					hash_id = row[0][0] #speicher das Ergebnis in Variable
					cur.execute("INSERT INTO enthält (tweet_id, hashtag_id) VALUES (%s, %s)", (i-1, hash_id)) #erstelle Datenbankeintrag für Tabelle enthält mit tweet_id und hash-id
					temp = '#' #speicher # in temp für nächstes hashtag
					continue #und mach weiter mit nächstem Zeichen
				else: #für alle anderen Fälle
					temp += hashs[j] #hänge an temp den aktuellen Buchstaben
				
		conn.commit() #committe, damit alle Änderungen sichtbar werden

		cur.close() #Schliße den cursor

	except (Exception, psycopg2.DatabaseError) as error:
		print(error)
	finally:
		if conn is not None:
			conn.close() #beende die connection
			print('Database connection closed.')

def main():

	connect()

if __name__ == '__main__':
	main()
