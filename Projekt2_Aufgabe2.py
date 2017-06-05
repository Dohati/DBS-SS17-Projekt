# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import re

gDB = load_workbook('american-election-tweets.xlsx') #Gegebende Datenbank wird geladen


wsgDB = gDB.active #Aktives Worksheet der gegebenen Datenbank


nDB = Workbook() #Workbook für neue Datenbank wird angelegt

wsnDB = nDB.active #Aktives Worksheet der neuen Datenbank

#Herausfiltern der benötigten Informationen, basierend auf dem Relationalen Modells.

z=0
row_count = wsgDB.max_row
for i in range(1,row_count+1):
	wsnDB['A'+str(i)] = wsgDB['A'+str(i)].value #Kopieren von "handle" in die neue Datenbank
	wsnDB['B'+str(i)] = wsgDB['B'+str(i)].value #Kopieren von "text" in die neue Datenbank
	wsnDB['C'+str(i)] = wsgDB['E'+str(i)].value #Kopieren von "time" in die neue Datenbank
	wsnDB['D'+str(i)] = wsgDB['H'+str(i)].value #Kopieren von "retweet_count" in die neue Datenbank
	wsnDB['E'+str(i)] = wsgDB['I'+str(i)].value #Kopieren von "favorite_count" in die neue Datenbank
#Herausfiltern der Hashtags aus dem Text
	hashs = '' #Alle hashtags eines textes werden hier gespeichert
	hashtag = '' #Hashtags werden hier einzeln erfasst
	temp = wsgDB['B' + str(i)].value #Abspeichern des textes in einer Variable
	for j in temp: #Gehen dem derzeitigen text durch (jeden einzeln) und speichern alle hashtags ab
		if (j == '#'): #Hashtag im text temp gefunden
			for a in range(z, len(temp)): #Sollte ein hashtag gefunden worden sein, gehen wir den nachfolgenden Text durch um es zu erfassen
				if (temp[a]==' ' or temp[a]=='\n'): 
				#Wenn nicht erwünschtes Zeichen gedunden
					break #...ende des Hashtags erreicht
				else:
					hashtag += temp[a] #Wenn kein space gefunden, dann ist a Teil des Hashtags
			if (len(hashtag) > 1):	#Wurde ein Hashtag falsch verwendet, wird es hier herausgefiltert
				hashs += re.sub("[^a-zA-Z0-9#]", "",hashtag).upper() #Speichere das gefundene Hashtag in hashs
		hashtag=''
		z+=1
	wsnDB['F' +str(i)] = hashs #Speichern der Hashtags in die entsprechende Zelle
	z=0
wsnDB['F1'] = 'hashtag' #Name der neuen Spalte
nDB.save('neueDatenbank.xlsx') #Speichern der neuen Datenbank
